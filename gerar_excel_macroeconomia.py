# -*- coding: utf-8 -*-
"""
================================================================================
 GERADOR DO BANCO DE DADOS MACROECONÔMICO DO BRASIL (2016-2025)
 Fontes: Banco Central do Brasil (SGS / Olinda), IBGE (via SGS), Ipeadata
================================================================================

USO (no seu computador, com internet):

    pip install requests pandas openpyxl
    python gerar_macroeconomia_brasil.py

Saída:  Macroeconomia_Brasil_2016_2025.xlsx

Modo template (sem internet, gera apenas a estrutura + metadados):

    python gerar_macroeconomia_brasil.py --template

Características:
- Coleta EXCLUSIVAMENTE de APIs oficiais (api.bcb.gov.br, olinda.bcb.gov.br,
  ipeadata.gov.br). Nenhum dado é inventado.
- Séries que falham na coleta são registradas na aba "Relatório de Coleta"
  (o script nunca aborta por causa de uma série).
- Cada série carrega fonte, código SGS e link na aba "Metadados".
- Códigos marcados como "verificar" devem ser conferidos no buscador do SGS
  (https://www3.bcb.gov.br/sgspub) antes de uso profissional definitivo.
================================================================================
"""

import argparse
import datetime as dt
import json
import sys
import time

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import requests
except ImportError:
    requests = None

# ------------------------------------------------------------------ parâmetros
DATA_INI = "01/01/2016"
DATA_FIM = "31/12/2025"
ANO_INI, ANO_FIM = 2016, 2025
ARQ_SAIDA = "Macroeconomia_Brasil_2016_2025.xlsx"
SGS_URL = ("https://api.bcb.gov.br/dados/serie/bcdata.sgs.{cod}/dados"
           "?formato=json&dataInicial={ini}&dataFinal={fim}")
SGS_BUSCA = "https://www3.bcb.gov.br/sgspub"

# ------------------------------------------------------------------ catálogo
# freq: 'D' diária, 'M' mensal, 'T' trimestral, 'A' anual
# agg_m: como converter diária->mensal ('last' | 'mean')
# agg_a: como gerar a série anual ('last' | 'mean' | 'sum' | 'acum' | None)
#        'acum' = capitalização composta de variações % mensais
# status: 'confirmado' | 'verificar' | 'manual'
S = lambda **kw: kw  # açúcar sintático

CATALOGO = {
    "PIB": [
        S(nome="PIB mensal - valores correntes", sgs=4380, unidade="R$ milhões",
          freq="M", agg_a="sum", status="confirmado",
          desc="PIB mensal estimado pelo BCB a preços correntes"),
        S(nome="PIB acumulado 12 meses - correntes", sgs=4382, unidade="R$ milhões",
          freq="M", agg_a="last", status="confirmado",
          desc="PIB acumulado nos últimos 12 meses, preços correntes"),
        S(nome="PIB - variação real anual", sgs=7326, unidade="% a.a.",
          freq="A", agg_a=None, status="confirmado",
          desc="Taxa de crescimento real do PIB (IBGE, Contas Nacionais)"),
        S(nome="PIB per capita / componentes (consumo, FBCF, exportações...)",
          sgs=None, unidade="-", freq="T", agg_a=None, status="manual",
          desc="Coletar no IBGE/SIDRA, Contas Nacionais Trimestrais "
               "(tabelas 1620, 1621, 6613, 6784) - https://sidra.ibge.gov.br"),
    ],
    "Inflação": [
        S(nome="IPCA - variação mensal", sgs=433, unidade="% a.m.", freq="M",
          agg_a="acum", status="confirmado", desc="Índice oficial de inflação (IBGE)"),
        S(nome="IPCA - acumulado 12 meses", sgs=13522, unidade="% 12m", freq="M",
          agg_a="last", status="confirmado", desc="IPCA acumulado em 12 meses"),
        S(nome="INPC - variação mensal", sgs=188, unidade="% a.m.", freq="M",
          agg_a="acum", status="confirmado", desc="Índice Nacional de Preços ao Consumidor (IBGE)"),
        S(nome="IGP-M - variação mensal", sgs=189, unidade="% a.m.", freq="M",
          agg_a="acum", status="confirmado", desc="Índice Geral de Preços - Mercado (FGV)"),
        S(nome="Meta de inflação", sgs=13521, unidade="% a.a.", freq="A",
          agg_a=None, status="confirmado", desc="Meta anual definida pelo CMN"),
    ],
    "Juros": [
        S(nome="Selic - meta definida pelo Copom", sgs=432, unidade="% a.a.",
          freq="D", agg_m="last", agg_a="last", status="confirmado",
          desc="Meta da taxa Selic definida pelo Copom"),
        S(nome="Selic efetiva - acumulada no mês", sgs=4390, unidade="% a.m.",
          freq="M", agg_a="acum", status="confirmado",
          desc="Taxa Selic efetiva acumulada no mês"),
        S(nome="Selic efetiva anualizada (base 252)", sgs=1178, unidade="% a.a.",
          freq="D", agg_m="last", agg_a="last", status="confirmado",
          desc="Taxa Selic diária anualizada, base 252 dias úteis"),
        S(nome="CDI - acumulado no mês", sgs=4391, unidade="% a.m.",
          freq="M", agg_a="acum", status="confirmado",
          desc="Taxa DI-Cetip acumulada no mês"),
        S(nome="Curva de juros (ETTJ)", sgs=None, unidade="-", freq="D",
          agg_a=None, status="manual",
          desc="Estrutura a termo: coletar na ANBIMA - https://www.anbima.com.br"),
    ],
    "Câmbio": [
        S(nome="Dólar PTAX venda (fim de mês)", sgs=1, unidade="R$/US$",
          freq="D", agg_m="last", agg_a="last", status="confirmado",
          desc="Taxa de câmbio livre, PTAX venda, último dia útil do mês"),
        S(nome="Dólar - taxa média mensal (venda)", sgs=3698, unidade="R$/US$",
          freq="M", agg_a="mean", status="confirmado",
          desc="Média mensal da PTAX venda"),
        S(nome="Euro PTAX venda (fim de mês)", sgs=21619, unidade="R$/EUR",
          freq="D", agg_m="last", agg_a="last", status="verificar",
          desc="Taxa de câmbio EUR/BRL, PTAX venda"),
        S(nome="Reservas internacionais (liquidez)", sgs=13621, unidade="US$ milhões",
          freq="M", agg_a="last", status="confirmado",
          desc="Reservas internacionais - conceito liquidez internacional"),
        S(nome="Posição de swap cambial", sgs=None, unidade="-", freq="M",
          agg_a=None, status="manual",
          desc="Coletar em Notas econômico-financeiras do BCB (setor externo)"),
    ],
    "Mercado de Trabalho": [
        S(nome="Taxa de desocupação (PNAD Contínua)", sgs=24369, unidade="%",
          freq="M", agg_a="mean", status="confirmado",
          desc="Trimestre móvel, IBGE/PNAD Contínua"),
        S(nome="População ocupada (PNAD Contínua)", sgs=24379, unidade="mil pessoas",
          freq="M", agg_a="mean", status="verificar",
          desc="Pessoas ocupadas, trimestre móvel"),
        S(nome="Rendimento médio real habitual", sgs=24382, unidade="R$",
          freq="M", agg_a="mean", status="verificar",
          desc="Rendimento médio real habitualmente recebido"),
        S(nome="Massa salarial real", sgs=None, unidade="-", freq="M",
          agg_a=None, status="manual",
          desc="Coletar no IBGE/SIDRA (PNAD Contínua) - massa de rendimento real"),
    ],
    "Fiscal": [
        S(nome="Dívida Bruta do Governo Geral", sgs=13762, unidade="% PIB",
          freq="M", agg_a="last", status="confirmado",
          desc="DBGG - metodologia BCB (2008)"),
        S(nome="Dívida Líquida do Setor Público", sgs=4513, unidade="% PIB",
          freq="M", agg_a="last", status="confirmado",
          desc="DLSP consolidada"),
        S(nome="Resultado primário - acum. 12m", sgs=5793, unidade="% PIB",
          freq="M", agg_a="last", status="verificar",
          desc="Setor público consolidado, acumulado em 12 meses"),
        S(nome="Resultado nominal - acum. 12m", sgs=5811, unidade="% PIB",
          freq="M", agg_a="last", status="verificar",
          desc="NFSP conceito nominal, acumulado em 12 meses"),
        S(nome="Carga tributária", sgs=None, unidade="% PIB", freq="A",
          agg_a=None, status="manual",
          desc="Receita Federal - Carga Tributária no Brasil (relatório anual)"),
    ],
    "Crédito": [
        S(nome="Saldo de crédito total", sgs=20539, unidade="R$ milhões",
          freq="M", agg_a="last", status="confirmado",
          desc="Saldo das operações de crédito do SFN - total"),
        S(nome="Saldo de crédito - Pessoa Jurídica", sgs=20540, unidade="R$ milhões",
          freq="M", agg_a="last", status="confirmado", desc="Saldo PJ"),
        S(nome="Saldo de crédito - Pessoa Física", sgs=20541, unidade="R$ milhões",
          freq="M", agg_a="last", status="confirmado", desc="Saldo PF"),
        S(nome="Taxa média de juros - total", sgs=20714, unidade="% a.a.",
          freq="M", agg_a="mean", status="verificar",
          desc="Taxa média das operações de crédito - total"),
        S(nome="Inadimplência - total", sgs=21082, unidade="%",
          freq="M", agg_a="last", status="verificar",
          desc="Atrasos superiores a 90 dias - carteira total"),
        S(nome="Spread médio - total", sgs=20783, unidade="p.p.",
          freq="M", agg_a="mean", status="verificar", desc="Spread médio do SFN"),
        S(nome="Crédito imobiliário / rural / consignado", sgs=None, unidade="-",
          freq="M", agg_a=None, status="manual",
          desc="Buscar no SGS por modalidade (menu Crédito) - " + SGS_BUSCA),
    ],
    "Sistema Financeiro": [
        S(nome="Base de dados PIX (quantidade e valores)", sgs=None, unidade="-",
          freq="M", agg_a=None, status="manual",
          desc="API Olinda Pix_DadosAbertos - o script tenta coletar automaticamente"),
        S(nome="Depósitos, poupança, TED, DOC, cartões, Open Finance, cooperativas",
          sgs=None, unidade="-", freq="M", agg_a=None, status="manual",
          desc="BCB - Estatísticas de pagamentos de varejo e SGS (menu Indicadores "
               "monetários e de crédito)"),
    ],
    "Setor Externo": [
        S(nome="Balança comercial - saldo mensal", sgs=22707, unidade="US$ milhões",
          freq="M", agg_a="sum", status="verificar",
          desc="Balanço de Pagamentos (BPM6) - saldo de bens"),
        S(nome="Transações correntes - mensal", sgs=22701, unidade="US$ milhões",
          freq="M", agg_a="sum", status="verificar",
          desc="Saldo em transações correntes (BPM6)"),
        S(nome="Investimento Direto no País (IDP)", sgs=22885, unidade="US$ milhões",
          freq="M", agg_a="sum", status="verificar",
          desc="Ingressos líquidos de investimento direto"),
        S(nome="Fluxo cambial total", sgs=13961, unidade="US$ milhões",
          freq="M", agg_a="sum", status="verificar",
          desc="Movimento de câmbio contratado - saldo"),
        S(nome="Reservas internacionais", sgs=13621, unidade="US$ milhões",
          freq="M", agg_a="last", status="confirmado", desc="Conceito liquidez"),
    ],
    "Agregados Monetários": [
        S(nome="Base monetária (média dias úteis)", sgs=1788, unidade="R$ milhões",
          freq="M", agg_a="last", status="verificar",
          desc="Base monetária restrita - média nos dias úteis"),
        S(nome="M1, M2, M3, M4", sgs=None, unidade="R$ milhões", freq="M",
          agg_a=None, status="manual",
          desc="Buscar no SGS: 'Meios de pagamento amplos' - " + SGS_BUSCA),
    ],
    "IBC-Br": [
        S(nome="IBC-Br - dessazonalizado", sgs=24363, unidade="Índice (2002=100)",
          freq="M", agg_a="mean", status="confirmado",
          desc="Índice de Atividade Econômica do BC, com ajuste sazonal"),
        S(nome="IBC-Br - observado", sgs=24364, unidade="Índice (2002=100)",
          freq="M", agg_a="mean", status="confirmado",
          desc="Índice de Atividade Econômica do BC, sem ajuste"),
    ],
    "Indicadores Financeiros": [
        S(nome="CDI - acumulado no mês", sgs=4391, unidade="% a.m.", freq="M",
          agg_a="acum", status="confirmado", desc="Taxa DI-Cetip"),
        S(nome="Risco-país (EMBI+ Brasil)", sgs=None, unidade="pontos-base",
          freq="D", agg_m="last", agg_a="last", status="confirmado",
          desc="J.P. Morgan via Ipeadata (JPM366_EMBI366)", especial="embi"),
        S(nome="Expectativas Focus (IPCA, Selic, PIB, Câmbio - ano corrente)",
          sgs=None, unidade="diversas", freq="M", agg_a=None,
          status="confirmado", desc="API Olinda/Expectativas - mediana anual",
          especial="focus"),
        S(nome="Inflação implícita (NTN-B)", sgs=None, unidade="% a.a.", freq="D",
          agg_a=None, status="manual", desc="ANBIMA - ETTJ / breakeven"),
    ],
}

FONTE_PADRAO = "Banco Central do Brasil - SGS"
LINK_SGS = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.{cod}/dados?formato=json"

# ------------------------------------------------------------------ coleta

def _sessao():
    s = requests.Session()
    s.headers.update({"User-Agent": "Mozilla/5.0 (coleta-dados-oficiais)"})
    return s


def fetch_sgs(sess, cod, freq):
    """Baixa uma série do SGS. Séries diárias são baixadas em janelas de 5 anos."""
    janelas = []
    if freq == "D":
        for a0 in range(ANO_INI, ANO_FIM + 1, 5):
            a1 = min(a0 + 4, ANO_FIM)
            janelas.append((f"01/01/{a0}", f"31/12/{a1}"))
    else:
        janelas.append((DATA_INI, DATA_FIM))
    registros = []
    for ini, fim in janelas:
        url = SGS_URL.format(cod=cod, ini=ini, fim=fim)
        r = sess.get(url, timeout=60)
        r.raise_for_status()
        registros.extend(r.json())
        time.sleep(0.4)
    if not registros:
        raise ValueError("série vazia")
    df = pd.DataFrame(registros)
    df["data"] = pd.to_datetime(df["data"], format="%d/%m/%Y")
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
    return df.dropna().set_index("data")["valor"].sort_index()


def fetch_embi(sess):
    url = ("http://www.ipeadata.gov.br/api/odata4/"
           "ValoresSerie(SERCODIGO='JPM366_EMBI366')")
    r = sess.get(url, timeout=60)
    r.raise_for_status()
    df = pd.DataFrame(r.json()["value"])
    df["data"] = pd.to_datetime(df["VALDATA"].str[:10])
    s = df.set_index("data")["VALVALOR"].astype(float).sort_index()
    return s[(s.index.year >= ANO_INI) & (s.index.year <= ANO_FIM)]


def fetch_focus(sess):
    """Mediana anual Focus (última pesquisa de cada mês, ano de referência corrente)."""
    base = ("https://olinda.bcb.gov.br/olinda/servico/Expectativas/versao/v1/"
            "odata/ExpectativasMercadoAnuais?$format=json&$top=100000"
            "&$filter=Indicador%20eq%20'{ind}'%20and%20Data%20ge%20'2016-01-01'")
    out = {}
    for ind, rot in [("IPCA", "Focus IPCA (mediana, ano corrente)"),
                     ("Selic", "Focus Selic (mediana, fim do ano corrente)"),
                     ("PIB Total", "Focus PIB (mediana, ano corrente)"),
                     ("Câmbio", "Focus Câmbio (mediana, fim do ano corrente)")]:
        r = sess.get(base.format(ind=ind.replace(" ", "%20")), timeout=120)
        r.raise_for_status()
        df = pd.DataFrame(r.json()["value"])
        if df.empty:
            continue
        df["Data"] = pd.to_datetime(df["Data"])
        df = df[df["DataReferencia"].astype(str) == df["Data"].dt.year.astype(str)]
        df["ym"] = df["Data"].dt.to_period("M")
        df = df.sort_values("Data").groupby("ym").last()
        s = df["Mediana"].astype(float)
        s.index = s.index.to_timestamp()
        out[rot] = s
    return out


def fetch_pix(sess):
    """Estatísticas de transações Pix (API Olinda). Melhor esforço."""
    url = ("https://olinda.bcb.gov.br/olinda/servico/Pix_DadosAbertos/versao/v1/"
           "odata/PixLiquidadosAtual?$format=json&$top=100000")
    r = sess.get(url, timeout=120)
    r.raise_for_status()
    df = pd.DataFrame(r.json()["value"])
    return df  # gravada em formato bruto na aba Sistema Financeiro


# ------------------------------------------------------------------ estilo
AZUL = "1F3864"      # azul-marinho (cabeçalhos)
AZUL_CLARO = "D9E2F3"
DOURADO = "BF9000"
FONTE = "Arial"

def _cab(ws, linha, textos, fill=AZUL, cor="FFFFFF"):
    for j, t in enumerate(textos, start=1):
        c = ws.cell(row=linha, column=j, value=t)
        c.font = Font(name=FONTE, bold=True, color=cor, size=10)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="thin"))

def _titulo(ws, cel, txt):
    ws[cel] = txt
    ws[cel].font = Font(name=FONTE, bold=True, size=13, color=AZUL)

def _fmt_num(ws, col, first_row, last_row, fmt="#,##0.00"):
    for r in range(first_row, last_row + 1):
        ws.cell(row=r, column=col).number_format = fmt
        ws.cell(row=r, column=col).font = Font(name=FONTE, size=10)

# ------------------------------------------------------------------ montagem

def escreve_tabela(ws, df, col_ini, titulo, com_filtro=False):
    """Escreve um DataFrame (índice = Período AAAA-MM ou ano) a partir de col_ini."""
    _titulo(ws, f"{get_column_letter(col_ini)}1", titulo)
    headers = ["Período"] + list(df.columns)
    for j, h in enumerate(headers):
        c = ws.cell(row=2, column=col_ini + j, value=h)
        c.font = Font(name=FONTE, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (idx, row) in enumerate(df.iterrows(), start=3):
        ws.cell(row=i, column=col_ini, value=str(idx)).font = Font(name=FONTE, size=10)
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=col_ini + j,
                        value=(None if pd.isna(v) else float(v)))
            c.number_format = "#,##0.00"
            c.font = Font(name=FONTE, size=10)
    for j in range(len(headers)):
        ws.column_dimensions[get_column_letter(col_ini + j)].width = 18 if j else 12
    if com_filtro and len(df):
        ws.auto_filter.ref = (f"{get_column_letter(col_ini)}2:"
                              f"{get_column_letter(col_ini + len(headers) - 1)}"
                              f"{2 + len(df)}")
    ws.freeze_panes = "A3"


def monta_aba_metadados(wb, coleta_log):
    ws = wb.create_sheet("Metadados")
    _titulo(ws, "A1", "Dicionário de Dados / Metadados")
    cab = ["Aba", "Nome", "Descrição", "Unidade", "Periodicidade", "Código SGS",
           "Fonte", "Link da fonte", "Última atualização", "Status do código"]
    _cab(ws, 2, cab)
    r = 3
    per = {"D": "Diária", "M": "Mensal", "T": "Trimestral", "A": "Anual"}
    for aba, series in CATALOGO.items():
        for s in series:
            link = (LINK_SGS.format(cod=s["sgs"]) if s.get("sgs") else
                    s.get("link", SGS_BUSCA))
            fonte = s.get("fonte", FONTE_PADRAO if s.get("sgs") else "ver descrição")
            ultima = coleta_log.get((aba, s["nome"]), {}).get("ultima", "-")
            vals = [aba, s["nome"], s["desc"], s["unidade"], per.get(s["freq"], s["freq"]),
                    s.get("sgs") or "-", fonte, link, ultima, s["status"]]
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = Font(name=FONTE, size=10)
                c.alignment = Alignment(vertical="top", wrap_text=(j in (3, 8)))
            if s["status"] != "confirmado":
                for j in range(1, len(vals) + 1):
                    ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor="FFF2CC")
            r += 1
    widths = [16, 38, 55, 14, 13, 11, 26, 45, 16, 14]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.auto_filter.ref = f"A2:J{r-1}"
    ws.freeze_panes = "A3"
    r += 1
    nota = ("NOTA: linhas em amarelo indicam código SGS a verificar no buscador oficial "
            f"({SGS_BUSCA}) ou coleta manual. Nenhum dado desta planilha foi estimado: "
            "todas as células provêm das APIs oficiais listadas acima.")
    ws.cell(row=r, column=1, value=nota).font = Font(name=FONTE, italic=True, size=9)


def monta_relatorio(wb, coleta_log, erros):
    ws = wb.create_sheet("Relatório de Coleta")
    _titulo(ws, "A1", "Relatório de Coleta e Controle de Qualidade")
    _cab(ws, 2, ["Aba", "Série", "Observações", "Última observação", "Status"])
    r = 3
    for (aba, nome), info in coleta_log.items():
        vals = [aba, nome, info.get("n", 0), info.get("ultima", "-"), "OK"]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=r, column=j, value=v).font = Font(name=FONTE, size=10)
        r += 1
    for aba, nome, msg in erros:
        vals = [aba, nome, msg, "-", "FALHA"]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(name=FONTE, size=10)
            c.fill = PatternFill("solid", fgColor="FCE4EC")
        r += 1
    for j, w in enumerate([16, 42, 30, 18, 10], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A3"


def monta_dashboard(wb, dados_m):
    ws = wb.create_sheet("Dashboard", 0)
    _titulo(ws, "A1", "Dashboard - Macroeconomia Brasil (2016-2025)")
    ws["A2"] = "Fontes: Banco Central do Brasil (SGS/Olinda), IBGE, Ipeadata"
    ws["A2"].font = Font(name=FONTE, italic=True, size=9)
    kpis = [
        ("PIB acum. 12m (R$ mi)", ("PIB", "PIB acumulado 12 meses - correntes")),
        ("IPCA 12 meses (%)", ("Inflação", "IPCA - acumulado 12 meses")),
        ("Selic meta (% a.a.)", ("Juros", "Selic - meta definida pelo Copom")),
        ("Dólar PTAX (R$/US$)", ("Câmbio", "Dólar PTAX venda (fim de mês)")),
        ("Desemprego (%)", ("Mercado de Trabalho", "Taxa de desocupação (PNAD Contínua)")),
        ("Crédito total (R$ mi)", ("Crédito", "Saldo de crédito total")),
        ("DBGG (% PIB)", ("Fiscal", "Dívida Bruta do Governo Geral")),
        ("Reservas (US$ mi)", ("Câmbio", "Reservas internacionais (liquidez)")),
        ("IBC-Br (dessaz.)", ("IBC-Br", "IBC-Br - dessazonalizado")),
    ]
    _cab(ws, 4, ["Indicador", "Último valor", "Referência"])
    r = 5
    for rotulo, chave in kpis:
        serie = dados_m.get(chave)
        c1 = ws.cell(row=r, column=1, value=rotulo)
        c1.font = Font(name=FONTE, size=10, bold=True)
        if serie is not None and len(serie):
            ws.cell(row=r, column=2, value=float(serie.dropna().iloc[-1]))
            ws.cell(row=r, column=3, value=str(serie.dropna().index[-1]))
        else:
            ws.cell(row=r, column=2, value="n/d")
        ws.cell(row=r, column=2).number_format = "#,##0.00"
        for j in (2, 3):
            ws.cell(row=r, column=j).font = Font(name=FONTE, size=10)
        r += 1
    for j, w in enumerate([26, 16, 12], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # mini-base para gráficos (colunas F em diante)
    graficos = [
        ("IPCA acumulado 12 meses (%)", ("Inflação", "IPCA - acumulado 12 meses")),
        ("Selic meta (% a.a.)", ("Juros", "Selic - meta definida pelo Copom")),
        ("Dólar PTAX venda (R$/US$)", ("Câmbio", "Dólar PTAX venda (fim de mês)")),
        ("Taxa de desocupação (%)", ("Mercado de Trabalho",
                                     "Taxa de desocupação (PNAD Contínua)")),
        ("DBGG (% PIB)", ("Fiscal", "Dívida Bruta do Governo Geral")),
        ("IBC-Br dessazonalizado", ("IBC-Br", "IBC-Br - dessazonalizado")),
    ]
    col0 = 6  # coluna F
    idx_ref = None
    for k, (_, chave) in enumerate(graficos):
        s = dados_m.get(chave)
        if s is None:
            continue
        if idx_ref is None:
            idx_ref = s.index
            ws.cell(row=1, column=col0, value="Período").font = Font(name=FONTE, bold=True, size=9)
            for i, p in enumerate(idx_ref, start=2):
                ws.cell(row=i, column=col0, value=str(p)).font = Font(name=FONTE, size=8)
        col = col0 + 1 + k
        ws.cell(row=1, column=col, value=graficos[k][0]).font = Font(name=FONTE, bold=True, size=9)
        s2 = s.reindex(idx_ref)
        for i, v in enumerate(s2, start=2):
            ws.cell(row=i, column=col,
                    value=(None if pd.isna(v) else float(v))).number_format = "#,##0.00"
    if idx_ref is not None:
        n = len(idx_ref)
        anchors = ["A16", "F16", "K16", "A32", "F32", "K32"]
        for k, (rot, _) in enumerate(graficos):
            ch = LineChart(); ch.title = rot; ch.style = 12
            ch.height, ch.width = 7.5, 12
            col = col0 + 1 + k
            ch.add_data(Reference(ws, min_col=col, min_row=1, max_row=n + 1),
                        titles_from_data=True)
            ch.set_categories(Reference(ws, min_col=col0, min_row=2, max_row=n + 1))
            ch.legend = None
            ws.add_chart(ch, anchors[k])
        for c in range(col0, col0 + len(graficos) + 1):
            ws.column_dimensions[get_column_letter(c)].hidden = True


def monta_resumo(wb, dados_m, template):
    ws = wb.create_sheet("Resumo Executivo")
    _titulo(ws, "A1", "Resumo Executivo")
    if template:
        ws["A3"] = ("Este resumo é gerado automaticamente quando o script é executado "
                    "com acesso à internet (valores de abertura, fechamento, mínimo, "
                    "máximo e média de cada indicador-chave no período 2016-2025).")
        ws["A3"].font = Font(name=FONTE, size=10, italic=True)
        return
    _cab(ws, 3, ["Indicador", "Início", "Fim", "Mínimo", "Máximo", "Média"])
    alvo = [("Inflação", "IPCA - acumulado 12 meses"),
            ("Juros", "Selic - meta definida pelo Copom"),
            ("Câmbio", "Dólar PTAX venda (fim de mês)"),
            ("Mercado de Trabalho", "Taxa de desocupação (PNAD Contínua)"),
            ("Fiscal", "Dívida Bruta do Governo Geral"),
            ("Crédito", "Saldo de crédito total"),
            ("Câmbio", "Reservas internacionais (liquidez)"),
            ("IBC-Br", "IBC-Br - dessazonalizado")]
    r = 4
    for chave in alvo:
        s = dados_m.get(chave)
        if s is None or not len(s):
            continue
        s = s.dropna()
        vals = [chave[1], float(s.iloc[0]), float(s.iloc[-1]),
                float(s.min()), float(s.max()), float(s.mean())]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(name=FONTE, size=10)
            if j > 1:
                c.number_format = "#,##0.00"
        r += 1
    for j, w in enumerate([40, 12, 12, 12, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def anualiza(s, modo):
    if modo is None:
        return None
    g = s.groupby(s.index.str[:4])
    if modo == "last":
        return g.last()
    if modo == "mean":
        return g.mean()
    if modo == "sum":
        return g.sum()
    if modo == "acum":  # composição de variações % mensais
        return g.apply(lambda x: ((1 + x / 100).prod() - 1) * 100)
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", action="store_true",
                    help="Gera apenas a estrutura (sem coleta de dados)")
    ap.add_argument("--saida", default=ARQ_SAIDA)
    args = ap.parse_args()

    template = args.template
    dados_m = {}   # (aba, nome) -> pd.Series índice 'AAAA-MM'
    coleta_log, erros = {}, []

    if not template:
        if requests is None:
            sys.exit("Instale as dependências: pip install requests pandas openpyxl")
        sess = _sessao()
        for aba, series in CATALOGO.items():
            for s in series:
                nome = s["nome"]
                try:
                    if s.get("especial") == "embi":
                        raw = fetch_embi(sess)
                    elif s.get("especial") == "focus":
                        for rot, sf in fetch_focus(sess).items():
                            sm = sf.copy(); sm.index = sm.index.strftime("%Y-%m")
                            dados_m[(aba, rot)] = sm
                            coleta_log[(aba, rot)] = {"n": len(sm),
                                                      "ultima": sm.index[-1]}
                        continue
                    elif s.get("sgs"):
                        raw = fetch_sgs(sess, s["sgs"], s["freq"])
                    else:
                        continue  # manual
                    if s["freq"] == "D":
                        raw = (raw.resample("ME").last() if s.get("agg_m", "last") == "last"
                               else raw.resample("ME").mean())
                    if s["freq"] == "A":
                        sm = raw.copy(); sm.index = sm.index.strftime("%Y")
                    else:
                        sm = raw.copy(); sm.index = sm.index.strftime("%Y-%m")
                    dados_m[(aba, nome)] = sm
                    coleta_log[(aba, nome)] = {"n": len(sm), "ultima": str(sm.index[-1])}
                    print(f"[OK]    {aba} | {nome} ({len(sm)} obs., última {sm.index[-1]})")
                except Exception as e:
                    erros.append((aba, nome, f"{type(e).__name__}: {e}"))
                    print(f"[FALHA] {aba} | {nome}: {e}")
        # Pix (melhor esforço)
        try:
            pix = fetch_pix(sess)
            dados_m[("__pix__", "__pix__")] = pix
            print(f"[OK]    Sistema Financeiro | Pix ({len(pix)} linhas)")
        except Exception as e:
            erros.append(("Sistema Financeiro", "Pix (Olinda)", str(e)))

    wb = Workbook(); wb.remove(wb.active)

    # ---- abas temáticas
    for aba, series in CATALOGO.items():
        ws = wb.create_sheet(aba)
        mensais = {s["nome"]: dados_m.get((aba, s["nome"]))
                   for s in series if s["freq"] in ("D", "M")}
        mensais = {k: v for k, v in mensais.items() if v is not None}
        anuais = {}
        for s in series:
            sm = dados_m.get((aba, s["nome"]))
            if sm is None:
                continue
            if s["freq"] == "A":
                anuais[s["nome"]] = sm
            elif s.get("agg_a"):
                sa = anualiza(sm, s["agg_a"])
                if sa is not None:
                    anuais[s["nome"]] = sa
        # Focus entra na aba Indicadores Financeiros
        if aba == "Indicadores Financeiros":
            for (a2, n2), v in dados_m.items():
                if a2 == aba and n2.startswith("Focus"):
                    mensais[n2] = v
        if template or not mensais:
            df_m = pd.DataFrame(columns=[s["nome"] for s in series
                                         if s["freq"] in ("D", "M") and
                                         (s.get("sgs") or s.get("especial"))])
            df_m.index.name = "Período"
        else:
            df_m = pd.DataFrame(mensais)
        escreve_tabela(ws, df_m, 1, f"{aba} - Séries Mensais (AAAA-MM)", com_filtro=True)
        col_a = max(3, len(df_m.columns) + 3)
        if anuais and not template:
            escreve_tabela(ws, pd.DataFrame(anuais), col_a, f"{aba} - Séries Anuais")
        else:
            _titulo(ws, f"{get_column_letter(col_a)}1", f"{aba} - Séries Anuais")
            ws.cell(row=2, column=col_a,
                    value="(preenchido automaticamente na execução com internet)"
                    ).font = Font(name=FONTE, italic=True, size=9)
        # Pix bruto
        if aba == "Sistema Financeiro" and ("__pix__", "__pix__") in dados_m:
            pix = dados_m[("__pix__", "__pix__")]
            start = 3 + len(df_m) + 3
            _titulo(ws, f"A{start}", "Pix - dados abertos (Olinda/BCB)")
            for j, hcol in enumerate(pix.columns, start=1):
                ws.cell(row=start + 1, column=j, value=hcol).font = Font(name=FONTE, bold=True, size=9)
            for i, (_, row) in enumerate(pix.iterrows(), start=start + 2):
                for j, v in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=(str(v) if isinstance(v, (dict, list)) else v))

    # ---- dashboard, resumo, metadados, relatório
    if not template:
        monta_dashboard(wb, dados_m)
    else:
        ws = wb.create_sheet("Dashboard", 0)
        _titulo(ws, "A1", "Dashboard - Macroeconomia Brasil (2016-2025)")
        ws["A3"] = ("Os indicadores e gráficos são gerados automaticamente na execução "
                    "do script com internet: PIB, Inflação, Selic, Dólar, Desemprego, "
                    "Crédito, Dívida Pública, Reservas e IBC-Br.")
        ws["A3"].font = Font(name=FONTE, size=10, italic=True)
    monta_resumo(wb, dados_m, template)
    monta_aba_metadados(wb, coleta_log)
    monta_relatorio(wb, coleta_log, erros)

    # ---- Leia-me
    ws = wb.create_sheet("Leia-me", 1)
    _titulo(ws, "A1", "Leia-me")
    linhas = [
        "Banco de dados macroeconômico do Brasil - 2016 a 2025.",
        "Todos os valores provêm de APIs oficiais: Banco Central (SGS e Olinda),",
        "IBGE (via SGS) e Ipeadata. Nenhum dado foi estimado ou inventado.",
        "",
        "Como atualizar: execute 'python gerar_macroeconomia_brasil.py' com internet.",
        "Abas temáticas: tabela mensal (esquerda, com filtros) e tabela anual (direita).",
        "Datas no formato AAAA-MM; valores em formato numérico.",
        "A aba Metadados é o dicionário de dados completo (códigos SGS, fontes, links).",
        "A aba 'Relatório de Coleta' registra sucesso/falha de cada série.",
        "Linhas amarelas em Metadados = código a verificar no buscador do SGS.",
        f"Gerado em: {dt.date.today().isoformat()}"
        + (" (MODO TEMPLATE - sem dados)" if template else ""),
    ]
    for i, t in enumerate(linhas, start=3):
        ws.cell(row=i, column=1, value=t).font = Font(name=FONTE, size=10)
    ws.column_dimensions["A"].width = 90

    wb.save(args.saida)
    print(f"\nArquivo gerado: {args.saida}")
    if erros:
        print(f"Séries com falha: {len(erros)} (ver aba 'Relatório de Coleta')")


if __name__ == "__main__":
    main()
